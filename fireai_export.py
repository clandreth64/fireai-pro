"""FireAI Pro — data-format exporters.

Writes the orchestrator's non-drawing deliverables (hydraulics JSON, BOM XLSX,
and an IFC model) to disk from the authoritative deterministic design. These were
previously listed as downloadable but never actually written, so the download
endpoint 404'd ("file isn't available"). Each writer returns the filename on
success or None on failure, so the caller can register only files that exist.

No hard dependency on ifcopenshell (not reliably installed); the IFC is emitted
as a minimal but valid IFC2x3 STEP file (sprinklers as located proxies).
"""

from __future__ import annotations
import json
import os
from datetime import datetime, timezone


# ════════════════════════════════════════════════════════════════════════════
#  Hydraulics JSON
# ════════════════════════════════════════════════════════════════════════════

def write_hydraulics_json(design: dict, out_dir: str,
                          filename: str = "hydraulics.json") -> str | None:
    try:
        d = design or {}
        payload = {
            "generated":            datetime.now(timezone.utc).isoformat(),
            "compliant":            d.get("compliant"),
            "static_pressure_psi":  d.get("static_pressure"),
            "residual_pressure_psi": d.get("residual_pressure"),
            "required_pressure_psi": d.get("required_pressure"),
            "pressure_delta_psi":   d.get("pressure_delta"),
            "flow_demand_gpm":      d.get("flow_demand"),
            "density_area":         d.get("density_area"),
            "demand_curve":         d.get("demand_curve"),
            "remote_area_calcs":    d.get("remote_area_calcs"),
            "fire_pump_added":      d.get("fire_pump_added", False),
            "fire_pump":            d.get("fire_pump"),
            "compliance_flags":     (d.get("design_metadata", {}) or {}).get("compliance_flags", []),
        }
        path = os.path.join(out_dir, filename)
        with open(path, "w", encoding="utf-8") as fh:
            json.dump(payload, fh, indent=2, default=str)
        return filename
    except Exception:
        return None


# ════════════════════════════════════════════════════════════════════════════
#  BOM XLSX
# ════════════════════════════════════════════════════════════════════════════

def write_bom_xlsx(design: dict, out_dir: str,
                   filename: str = "bill_of_materials.xlsx") -> str | None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        bom = (design or {}).get("bom", []) or []
        wb = Workbook()
        ws = wb.active
        ws.title = "Bill of Materials"

        headers = ["Item", "Part Number", "Qty", "Unit", "Unit Cost ($)",
                   "Extended Cost ($)", "NFPA Ref"]
        ws.append(headers)
        hdr_font = Font(bold=True, color="FFFFFF")
        hdr_fill = PatternFill("solid", fgColor="1F3864")
        for c in ws[1]:
            c.font = hdr_font
            c.fill = hdr_fill
            c.alignment = Alignment(horizontal="center")

        total = 0.0
        for item in bom:
            qty  = float(item.get("qty", 0) or 0)
            cost = float(item.get("unit_cost", 0) or 0)
            ext  = round(qty * cost, 2)
            total += ext
            ws.append([
                item.get("item", item.get("description", "")),
                item.get("part_number", ""),
                qty,
                item.get("unit", "EA"),
                cost,
                ext,
                item.get("nfpa_ref", ""),
            ])

        ws.append([])
        ws.append(["", "", "", "", "TOTAL", round(total, 2), ""])
        ws[ws.max_row][4].font = Font(bold=True)
        ws[ws.max_row][5].font = Font(bold=True)

        widths = [46, 16, 8, 8, 14, 16, 18]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[chr(64 + i)].width = w

        path = os.path.join(out_dir, filename)
        wb.save(path)
        return filename
    except Exception:
        return None


# ════════════════════════════════════════════════════════════════════════════
#  IFC model (minimal valid IFC2x3 STEP file — no ifcopenshell dependency)
# ════════════════════════════════════════════════════════════════════════════

def write_model_ifc(design: dict, project: dict, out_dir: str,
                    filename: str = "model.ifc") -> str | None:
    try:
        d = design or {}
        proj = project or {}
        spk = d.get("sprinkler_placements", []) or []

        ts = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%S")
        proj_name = str(proj.get("project_name", "FireAI Project")).replace("'", "")
        site_name = str(proj.get("location", proj.get("project_address", "Site"))).replace("'", "")

        lines: list[str] = []
        _id = [0]

        def nid() -> int:
            _id[0] += 1
            return _id[0]

        def add(s: str) -> int:
            i = nid()
            lines.append(f"#{i}={s};")
            return i

        # ── Owner / context boilerplate ──────────────────────────────────────
        person   = add("IFCPERSON($,'','FireAI',$,$,$,$,$)")
        org      = add("IFCORGANIZATION($,'FireAI Pro',$,$,$)")
        pando    = add(f"IFCPERSONANDORGANIZATION(#{person},#{org},$)")
        app      = add(f"IFCAPPLICATION(#{org},'1.0','FireAI Pro','FIREAI')")
        owner    = add(f"IFCOWNERHISTORY(#{pando},#{app},$,.ADDED.,$,$,$,0)")

        dim      = add("IFCDIMENSIONALEXPONENTS(0,0,0,0,0,0,0)")
        len_unit = add("IFCSIUNIT(*,.LENGTHUNIT.,.MILLI.,.METRE.)")
        area_un  = add("IFCSIUNIT(*,.AREAUNIT.,$,.SQUARE_METRE.)")
        vol_un   = add("IFCSIUNIT(*,.VOLUMEUNIT.,$,.CUBIC_METRE.)")
        units    = add(f"IFCUNITASSIGNMENT((#{len_unit},#{area_un},#{vol_un}))")

        origin   = add("IFCCARTESIANPOINT((0.,0.,0.))")
        zdir     = add("IFCDIRECTION((0.,0.,1.))")
        xdir     = add("IFCDIRECTION((1.,0.,0.))")
        axis     = add(f"IFCAXIS2PLACEMENT3D(#{origin},#{zdir},#{xdir})")
        wcs      = add(f"IFCGEOMETRICREPRESENTATIONCONTEXT($,'Model',3,1.E-05,#{axis},$)")
        ctx      = add(f"IFCPROJECT('{_guid()}',#{owner},'{proj_name}',$,$,$,$,(#{wcs}),#{units})")

        site_pl  = add(f"IFCLOCALPLACEMENT($,#{axis})")
        site     = add(f"IFCSITE('{_guid()}',#{owner},'{site_name}',$,$,#{site_pl},$,$,.ELEMENT.,$,$,$,$,$)")
        bldg_pl  = add(f"IFCLOCALPLACEMENT(#{site_pl},#{axis})")
        bldg     = add(f"IFCBUILDING('{_guid()}',#{owner},'Building',$,$,#{bldg_pl},$,$,.ELEMENT.,$,$,$)")
        sto_pl   = add(f"IFCLOCALPLACEMENT(#{bldg_pl},#{axis})")
        storey   = add(f"IFCBUILDINGSTOREY('{_guid()}',#{owner},'Level 1',$,$,#{sto_pl},$,$,.ELEMENT.,0.)")

        add(f"IFCRELAGGREGATES('{_guid()}',#{owner},$,$,#{ctx},(#{site}))")
        add(f"IFCRELAGGREGATES('{_guid()}',#{owner},$,$,#{site},(#{bldg}))")
        add(f"IFCRELAGGREGATES('{_guid()}',#{owner},$,$,#{bldg},(#{storey}))")

        # ── Sprinklers as located proxies ────────────────────────────────────
        FT_TO_MM = 304.8
        proxies: list[int] = []
        for idx, s in enumerate(spk):
            x = float(s.get("x", 0)) * FT_TO_MM
            y = float(s.get("y", 0)) * FT_TO_MM
            z = float(s.get("z", s.get("elevation", 0)) or 0) * FT_TO_MM
            pt   = add(f"IFCCARTESIANPOINT(({x:.1f},{y:.1f},{z:.1f}))")
            ax   = add(f"IFCAXIS2PLACEMENT3D(#{pt},$,$)")
            plc  = add(f"IFCLOCALPLACEMENT(#{sto_pl},#{ax})")
            tag  = str(s.get("id", f"S-{idx+1:03d}"))
            prx  = add(f"IFCBUILDINGELEMENTPROXY('{_guid()}',#{owner},'Sprinkler {tag}',"
                       f"'{s.get('type','pendent')} K={s.get('k_factor','')}',$,#{plc},$,$,.NOTDEFINED.)")
            proxies.append(prx)

        if proxies:
            refs = ",".join(f"#{p}" for p in proxies)
            add(f"IFCRELCONTAINEDINSPATIALSTRUCTURE('{_guid()}',#{owner},$,$,({refs}),#{storey})")

        # ── STEP file assembly ───────────────────────────────────────────────
        body = "\n".join(lines)
        header = (
            "ISO-10303-21;\n"
            "HEADER;\n"
            f"FILE_DESCRIPTION(('ViewDefinition [CoordinationView]'),'2;1');\n"
            f"FILE_NAME('{filename}','{ts}',(''),(''),'FireAI Pro','FireAI Pro','');\n"
            "FILE_SCHEMA(('IFC2X3'));\n"
            "ENDSEC;\n"
            "DATA;\n"
        )
        footer = "\nENDSEC;\nEND-ISO-10303-21;\n"

        path = os.path.join(out_dir, filename)
        with open(path, "w", encoding="utf-8") as fh:
            fh.write(header + body + footer)
        return filename
    except Exception:
        return None


def _guid() -> str:
    """IFC-compressed 22-char GUID."""
    import uuid
    chars = "0123456789ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz_$"
    n = uuid.uuid4().int
    out = []
    for _ in range(22):
        out.append(chars[n & 63])
        n >>= 6
    return "".join(reversed(out))


# ════════════════════════════════════════════════════════════════════════════
#  Convenience: write whichever data files were requested, return real names
# ════════════════════════════════════════════════════════════════════════════

def write_requested(design: dict, project: dict, out_dir: str,
                    requested_filenames) -> list[str]:
    """For each requested data filename, write the real file; return those that
    succeeded. Drawing files (handled by the drawing engine) are ignored here."""
    os.makedirs(out_dir, exist_ok=True)
    written: list[str] = []
    want = set(requested_filenames or [])
    for fn in want:
        low = fn.lower()
        ok = None
        if low.endswith(".json"):
            ok = write_hydraulics_json(design, out_dir, fn)
        elif low.endswith(".xlsx"):
            ok = write_bom_xlsx(design, out_dir, fn)
        elif low.endswith(".ifc"):
            ok = write_model_ifc(design, project, out_dir, fn)
        if ok:
            written.append(ok)
    return written
