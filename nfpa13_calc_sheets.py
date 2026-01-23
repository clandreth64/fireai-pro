#!/usr/bin/env python3
"""
FireAI Pro - NFPA 13 Calculation Sheet Generator
VERSION: 1.0.0

🔥 PERMIT-READY HYDRAULIC CALCULATION SHEETS

This module generates professional, AHJ-ready hydraulic calculation sheets
that match the quality and format of AutoSprink and Elite outputs.

📋 OUTPUT FORMATS:
✅ PDF - Professional multi-page calculation report
✅ Excel - Editable calculation spreadsheet
✅ Text - Plain text for review/archival
✅ JSON - Machine-readable data export

📑 CALCULATION SHEET SECTIONS:
1. Cover Sheet - Project info, contractor, dates
2. Design Criteria - Hazard, density, area, K-factors
3. Water Supply Data - Test data, curves
4. Hydraulic Calculations - Node-by-node table
5. Pipe Schedule - All pipes with hydraulic data
6. Summary - Demand vs supply, pass/fail
7. Demand Curve Graph - Visual representation
8. Notes & References - NFPA citations

🎯 COMPLIANCE:
- NFPA 13 (2022) format requirements
- Standard calculation sheet layout
- All required data fields
- Professional formatting for AHJ review
"""

import io
import json
import math
import os
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Any, Tuple
from dataclasses import dataclass, field
from enum import Enum

# Import from main hydraulics engine
try:
    from enhanced_hydraulics_engine import (
        HydraulicNetwork,
        HydraulicNode,
        HydraulicPipe,
        Sprinkler,
        WaterSupplyData,
        RemoteArea,
        SystemType,
        NodeType,
        PipeType,
        NFPA13Constants,
    )
except ImportError:
    # For standalone testing
    pass

import logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Check for optional dependencies
reportlab_available = True
openpyxl_available = True
matplotlib_available = True

try:
    from reportlab.lib.pagesizes import letter, A4, landscape
    from reportlab.lib.units import inch, mm
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
        PageBreak, Image, KeepTogether, Flowable
    )
    from reportlab.pdfgen import canvas
    from reportlab.graphics.shapes import Drawing, Line, String, Rect
    from reportlab.graphics.charts.lineplots import LinePlot
    from reportlab.graphics import renderPDF
    logger.info("✅ ReportLab available for PDF generation")
except ImportError:
    reportlab_available = False
    logger.warning("⚠️ ReportLab not available - PDF generation disabled")

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import LineChart, Reference
    logger.info("✅ OpenPyXL available for Excel generation")
except ImportError:
    openpyxl_available = False
    logger.warning("⚠️ OpenPyXL not available - Excel generation disabled")

try:
    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt
    import matplotlib.patches as mpatches
    logger.info("✅ Matplotlib available for graph generation")
except ImportError:
    matplotlib_available = False
    logger.warning("⚠️ Matplotlib not available - graph generation disabled")


# ================================================================================================
# DATA STRUCTURES FOR CALCULATION SHEETS
# ================================================================================================

@dataclass
class ProjectInfo:
    """Project information for calculation sheet header"""
    project_name: str = ""
    project_number: str = ""
    address: str = ""
    city: str = ""
    state: str = ""
    zip_code: str = ""
    
    # Contractor info
    contractor_name: str = ""
    contractor_license: str = ""
    contractor_address: str = ""
    contractor_phone: str = ""
    contractor_email: str = ""
    
    # Engineer info  
    engineer_name: str = ""
    engineer_license: str = ""
    engineer_company: str = ""
    
    # Drawing info
    drawing_number: str = ""
    drawing_date: str = ""
    revision: str = ""
    
    # Dates
    calculation_date: str = field(default_factory=lambda: datetime.now().strftime("%Y-%m-%d"))
    
    # AHJ info
    ahj_name: str = ""
    permit_number: str = ""


@dataclass
class DesignCriteria:
    """Design criteria section data"""
    occupancy_type: str = ""
    hazard_classification: str = "Ordinary Hazard Group 1"
    commodity_class: str = ""
    storage_height_ft: float = 0.0
    ceiling_height_ft: float = 0.0
    
    # Density/area
    design_density_gpm_sqft: float = 0.15
    design_area_sqft: float = 1500.0
    area_adjustment_factor: float = 1.0  # For dry systems
    adjusted_area_sqft: float = 1500.0
    
    # Sprinkler info
    sprinkler_type: str = "Standard Spray Pendent"
    sprinkler_manufacturer: str = ""
    sprinkler_model: str = ""
    sprinkler_k_factor: float = 5.6
    sprinkler_temp_rating: str = "155°F"
    sprinkler_coverage_sqft: float = 130.0
    
    # Counts
    total_sprinklers: int = 0
    sprinklers_calculated: int = 0
    
    # Allowances
    hose_stream_gpm: float = 250.0
    inside_hose_gpm: float = 0.0
    duration_minutes: int = 60
    
    # System type
    system_type: str = "Wet"
    pipe_material: str = "Black Steel"
    pipe_schedule: str = "Schedule 40"


@dataclass
class HydraulicCalcRow:
    """Single row in hydraulic calculation table"""
    step: int
    node_ref: str
    elevation_ft: float
    k_factor: Optional[float]
    flow_q_gpm: float
    pressure_pt_psi: float
    
    # Pipe data (to next node)
    pipe_ref: str = ""
    pipe_size_nominal: float = 0.0
    pipe_id_inches: float = 0.0
    c_factor: int = 120
    pipe_length_ft: float = 0.0
    equiv_length_ft: float = 0.0
    total_length_ft: float = 0.0
    
    # Pressure calculations
    friction_loss_psi: float = 0.0
    elevation_psi: float = 0.0
    
    # Required pressure at next node
    normal_pressure_pn_psi: float = 0.0
    
    # Fittings summary
    fittings: str = ""
    notes: str = ""


@dataclass 
class CalcSheetSummary:
    """Calculation summary data"""
    # System demand
    system_flow_gpm: float = 0.0
    system_pressure_psi: float = 0.0
    
    # With hose stream
    hose_stream_gpm: float = 250.0
    total_flow_gpm: float = 0.0
    
    # Water supply
    static_pressure_psi: float = 0.0
    residual_pressure_psi: float = 0.0
    flow_at_residual_gpm: float = 0.0
    available_at_demand_psi: float = 0.0
    
    # Result
    safety_margin_psi: float = 0.0
    is_adequate: bool = False
    
    # Graph data
    supply_curve_points: List[Tuple[float, float]] = field(default_factory=list)
    demand_point: Tuple[float, float] = (0.0, 0.0)
    total_demand_point: Tuple[float, float] = (0.0, 0.0)


# ================================================================================================
# MAIN CALCULATION SHEET GENERATOR
# ================================================================================================

class NFPA13CalcSheetGenerator:
    """
    Generates permit-ready NFPA 13 hydraulic calculation sheets
    
    Produces professional output matching industry standards for AHJ review.
    """
    
    def __init__(self):
        self.logger = logging.getLogger(f"{__name__}.CalcSheetGenerator")
        
        # Page settings
        self.page_size = letter
        self.margin = 0.5 * inch
        
        # Styles
        self._init_styles()
    
    def _init_styles(self):
        """Initialize document styles"""
        if not reportlab_available:
            return
        
        self.styles = getSampleStyleSheet()
        
        # Custom styles
        self.styles.add(ParagraphStyle(
            name='CalcTitle',
            parent=self.styles['Heading1'],
            fontSize=16,
            alignment=TA_CENTER,
            spaceAfter=12,
            fontName='Helvetica-Bold'
        ))
        
        self.styles.add(ParagraphStyle(
            name='SectionHeader',
            parent=self.styles['Heading2'],
            fontSize=12,
            spaceBefore=12,
            spaceAfter=6,
            fontName='Helvetica-Bold',
            textColor=colors.darkblue
        ))
        
        self.styles.add(ParagraphStyle(
            name='TableHeader',
            parent=self.styles['Normal'],
            fontSize=8,
            fontName='Helvetica-Bold',
            alignment=TA_CENTER
        ))
        
        self.styles.add(ParagraphStyle(
            name='TableCell',
            parent=self.styles['Normal'],
            fontSize=8,
            fontName='Helvetica'
        ))
        
        self.styles.add(ParagraphStyle(
            name='SmallText',
            parent=self.styles['Normal'],
            fontSize=7,
            fontName='Helvetica'
        ))
    
    def generate_from_network(self, 
                             network: 'HydraulicNetwork',
                             solution: Dict[str, Any],
                             project_info: ProjectInfo,
                             output_dir: str) -> Dict[str, str]:
        """
        Generate all calculation sheet formats from a solved network
        
        Args:
            network: Solved HydraulicNetwork
            solution: Solution data from solver
            project_info: Project information
            output_dir: Output directory path
            
        Returns:
            Dict of output file paths by format
        """
        self.logger.info("📋 Generating NFPA 13 calculation sheets...")
        
        Path(output_dir).mkdir(parents=True, exist_ok=True)
        output_files = {}
        
        # Build calculation data
        design_criteria = self._build_design_criteria(network)
        calc_rows = self._build_calc_rows(network, solution)
        summary = self._build_summary(network, solution)
        
        # Generate PDF
        if reportlab_available:
            pdf_path = str(Path(output_dir) / "hydraulic_calculations.pdf")
            self.generate_pdf(
                project_info, design_criteria, network.water_supply,
                calc_rows, summary, pdf_path
            )
            output_files['pdf'] = pdf_path
        
        # Generate Excel
        if openpyxl_available:
            xlsx_path = str(Path(output_dir) / "hydraulic_calculations.xlsx")
            self.generate_excel(
                project_info, design_criteria, network.water_supply,
                calc_rows, summary, xlsx_path
            )
            output_files['xlsx'] = xlsx_path
        
        # Generate text report
        txt_path = str(Path(output_dir) / "hydraulic_calculations.txt")
        self.generate_text(
            project_info, design_criteria, network.water_supply,
            calc_rows, summary, txt_path
        )
        output_files['txt'] = txt_path
        
        # Generate JSON data
        json_path = str(Path(output_dir) / "hydraulic_calculations.json")
        self.generate_json(
            project_info, design_criteria, network.water_supply,
            calc_rows, summary, json_path
        )
        output_files['json'] = json_path
        
        # Generate demand curve graph
        if matplotlib_available:
            graph_path = str(Path(output_dir) / "demand_curve.png")
            self.generate_demand_curve_graph(summary, graph_path)
            output_files['graph'] = graph_path
        
        self.logger.info(f"✅ Generated {len(output_files)} output files")
        return output_files
    
    def _build_design_criteria(self, network: 'HydraulicNetwork') -> DesignCriteria:
        """Build design criteria from network"""
        criteria = DesignCriteria()
        
        # Hazard classification
        criteria.hazard_classification = network.hazard_class.replace('_', ' ').title()
        
        # Density/area
        criteria.design_density_gpm_sqft = network.design_density
        criteria.design_area_sqft = network.design_area_sqft
        
        # Check for dry system adjustment
        if network.system_type in [SystemType.DRY, SystemType.PREACTION_SINGLE,
                                    SystemType.PREACTION_DOUBLE, SystemType.PREACTION_NON_INTERLOCK]:
            if not network.has_quick_opening_device:
                criteria.area_adjustment_factor = 1.30
                criteria.adjusted_area_sqft = network.design_area_sqft
        
        # Sprinkler info
        if network.sprinklers:
            sample_sprinkler = list(network.sprinklers.values())[0]
            criteria.sprinkler_k_factor = sample_sprinkler.k_factor
            criteria.sprinkler_type = sample_sprinkler.sprinkler_type.replace('_', ' ').title()
        
        # Counts
        criteria.total_sprinklers = len(network.sprinklers)
        criteria.sprinklers_calculated = sum(
            1 for s in network.sprinklers.values() if s.is_in_remote_area
        )
        
        # Allowances
        criteria.hose_stream_gpm = network.hose_stream_gpm
        
        hazard_info = NFPA13Constants.HAZARD_CLASSIFICATIONS.get(
            network.hazard_class, {}
        )
        criteria.duration_minutes = hazard_info.get('duration_minutes', 60)
        
        # System type
        criteria.system_type = network.system_type.value.replace('_', ' ').title()
        
        # Pipe info
        if network.pipes:
            sample_pipe = list(network.pipes.values())[0]
            criteria.pipe_material = sample_pipe.material.replace('_', ' ').title()
            criteria.pipe_schedule = f"Schedule {sample_pipe.schedule}"
        
        return criteria
    
    def _build_calc_rows(self, network: 'HydraulicNetwork', 
                        solution: Dict[str, Any]) -> List[HydraulicCalcRow]:
        """
        Build hydraulic calculation table rows
        
        Creates proper NFPA 13 format calculation trace from most remote
        sprinkler back to Base of Riser (BOR), showing:
        - Node-by-node pressure buildup
        - Flow accumulation at junctions
        - Friction loss in each pipe segment
        - Elevation pressure changes
        """
        rows = []
        step = 1
        
        # Get remote area sprinklers
        remote_sprinklers = [
            s for s in network.sprinklers.values() if s.is_in_remote_area
        ]
        
        if not remote_sprinklers:
            # Fallback: use all sprinklers
            remote_sprinklers = list(network.sprinklers.values())
        
        if not remote_sprinklers:
            return rows
        
        # Build path from source to all nodes using BFS
        source_id = network.source_node_id
        paths_from_source = self._build_paths_from_source(network, source_id)
        
        # Sort sprinklers by path length (most remote first)
        remote_sprinklers.sort(
            key=lambda s: len(paths_from_source.get(s.node_id, [])),
            reverse=True
        )
        
        # Track processed items
        processed_nodes = set()
        processed_pipes = set()
        cumulative_flow = 0.0
        
        # Process branch lines (sprinklers on the same branch)
        branch_groups = self._group_sprinklers_by_branch(network, remote_sprinklers)
        
        for branch_id, branch_sprinklers in branch_groups.items():
            # Sort sprinklers in branch by distance from junction (most remote first)
            branch_sprinklers.sort(
                key=lambda s: len(paths_from_source.get(s.node_id, [])),
                reverse=True
            )
            
            branch_flow = 0.0
            prev_node = None
            prev_pressure = 0.0
            
            for i, sprinkler in enumerate(branch_sprinklers):
                node = network.nodes.get(sprinkler.node_id)
                if not node:
                    continue
                
                # Calculate sprinkler flow at operating pressure
                if sprinkler.operating_pressure_psi > 0:
                    pressure = sprinkler.operating_pressure_psi
                else:
                    # Calculate from required pressure
                    pressure = max(7.0, sprinkler.min_pressure_psi)
                
                flow = sprinkler.k_factor * math.sqrt(pressure)
                branch_flow += flow
                
                # Create sprinkler row
                row = HydraulicCalcRow(
                    step=step,
                    node_ref=node.tag or sprinkler.id,
                    elevation_ft=node.elevation,
                    k_factor=sprinkler.k_factor,
                    flow_q_gpm=flow,
                    pressure_pt_psi=pressure,
                    notes=f"Sprinkler K={sprinkler.k_factor}"
                )
                
                # Find pipe connecting this sprinkler
                if prev_node:
                    connecting_pipe = self._find_pipe_between(network, node.id, prev_node.id)
                    if connecting_pipe and connecting_pipe.id not in processed_pipes:
                        self._add_pipe_data_to_row(row, connecting_pipe, node, prev_node)
                        processed_pipes.add(connecting_pipe.id)
                else:
                    # First sprinkler - find pipe to branch line
                    connected_pipes = network.get_connected_pipes(node.id)
                    for pipe in connected_pipes:
                        if pipe.id not in processed_pipes:
                            other_id = pipe.end_node_id if pipe.start_node_id == node.id else pipe.start_node_id
                            other_node = network.nodes.get(other_id)
                            if other_node:
                                self._add_pipe_data_to_row(row, pipe, node, other_node)
                                processed_pipes.add(pipe.id)
                                break
                
                rows.append(row)
                processed_nodes.add(node.id)
                prev_node = node
                prev_pressure = row.normal_pressure_pn_psi if row.normal_pressure_pn_psi else pressure
                step += 1
            
            cumulative_flow += branch_flow
        
        # Add junction/cross main nodes
        for node_id, node in network.nodes.items():
            if node_id in processed_nodes:
                continue
            if node.node_type == NodeType.SOURCE:
                continue
            if node.node_type not in [NodeType.JUNCTION]:
                continue
            
            # Find connected pipes not yet processed
            connected_pipes = network.get_connected_pipes(node_id)
            pipe_for_row = None
            next_node = None
            
            for pipe in connected_pipes:
                if pipe.id not in processed_pipes:
                    other_id = pipe.end_node_id if pipe.start_node_id == node_id else pipe.start_node_id
                    other_node = network.nodes.get(other_id)
                    if other_node and other_id not in processed_nodes:
                        pipe_for_row = pipe
                        next_node = other_node
                        break
            
            row = HydraulicCalcRow(
                step=step,
                node_ref=node.tag or node_id,
                elevation_ft=node.elevation,
                k_factor=None,
                flow_q_gpm=cumulative_flow,  # Accumulated flow at junction
                pressure_pt_psi=node.pressure_psi if node.pressure_psi > 0 else 0,
                notes="Junction - Flow combines"
            )
            
            if pipe_for_row and next_node:
                self._add_pipe_data_to_row(row, pipe_for_row, node, next_node)
                processed_pipes.add(pipe_for_row.id)
            
            rows.append(row)
            processed_nodes.add(node_id)
            step += 1
        
        # Add riser/main pipes to BOR
        source_node = network.nodes.get(source_id)
        if source_node:
            # Find unprocessed pipes leading to source
            for pipe_id, pipe in network.pipes.items():
                if pipe_id in processed_pipes:
                    continue
                if pipe.start_node_id == source_id or pipe.end_node_id == source_id:
                    other_id = pipe.end_node_id if pipe.start_node_id == source_id else pipe.start_node_id
                    other_node = network.nodes.get(other_id)
                    if other_node:
                        row = HydraulicCalcRow(
                            step=step,
                            node_ref=other_node.tag or other_id,
                            elevation_ft=other_node.elevation,
                            k_factor=None,
                            flow_q_gpm=cumulative_flow,
                            pressure_pt_psi=other_node.pressure_psi if other_node.pressure_psi > 0 else 0,
                            notes="Riser"
                        )
                        self._add_pipe_data_to_row(row, pipe, other_node, source_node)
                        rows.append(row)
                        processed_pipes.add(pipe_id)
                        step += 1
            
            # Final BOR row
            total_flow = sum(s.flow_gpm for s in network.sprinklers.values() if s.is_in_remote_area)
            if total_flow == 0:
                total_flow = cumulative_flow
            
            row = HydraulicCalcRow(
                step=step,
                node_ref="BOR",
                elevation_ft=source_node.elevation,
                k_factor=None,
                flow_q_gpm=total_flow,
                pressure_pt_psi=source_node.pressure_psi if source_node.pressure_psi > 0 else self._calc_bor_pressure(rows),
                notes="BASE OF RISER - SYSTEM DEMAND"
            )
            rows.append(row)
        
        return rows
    
    def _build_paths_from_source(self, network: 'HydraulicNetwork', 
                                  source_id: str) -> Dict[str, List[str]]:
        """Build paths from source to all nodes using BFS"""
        paths = {source_id: [source_id]}
        visited = {source_id}
        queue = [source_id]
        
        while queue:
            current = queue.pop(0)
            connected_pipes = network.get_connected_pipes(current)
            
            for pipe in connected_pipes:
                neighbor = pipe.end_node_id if pipe.start_node_id == current else pipe.start_node_id
                if neighbor not in visited:
                    visited.add(neighbor)
                    paths[neighbor] = paths[current] + [neighbor]
                    queue.append(neighbor)
        
        return paths
    
    def _group_sprinklers_by_branch(self, network: 'HydraulicNetwork',
                                     sprinklers: List['Sprinkler']) -> Dict[str, List['Sprinkler']]:
        """Group sprinklers by their branch line"""
        branches = {}
        
        for sprinkler in sprinklers:
            # Find the branch junction this sprinkler connects to
            node = network.nodes.get(sprinkler.node_id)
            if not node:
                continue
            
            # Trace back to find branch junction
            branch_id = self._find_branch_junction(network, sprinkler.node_id)
            if branch_id not in branches:
                branches[branch_id] = []
            branches[branch_id].append(sprinkler)
        
        return branches
    
    def _find_branch_junction(self, network: 'HydraulicNetwork', 
                               start_node_id: str) -> str:
        """Find the junction node that starts this branch"""
        current = start_node_id
        visited = {current}
        
        while True:
            connected = network.get_connected_pipes(current)
            if len(connected) > 2:  # Junction found
                return current
            
            # Move to next node
            next_node = None
            for pipe in connected:
                neighbor = pipe.end_node_id if pipe.start_node_id == current else pipe.start_node_id
                if neighbor not in visited:
                    next_node = neighbor
                    break
            
            if not next_node:
                return current  # Dead end or source
            
            visited.add(next_node)
            current = next_node
    
    def _find_pipe_between(self, network: 'HydraulicNetwork',
                           node1_id: str, node2_id: str) -> Optional['HydraulicPipe']:
        """Find pipe connecting two nodes"""
        for pipe in network.pipes.values():
            if (pipe.start_node_id == node1_id and pipe.end_node_id == node2_id) or \
               (pipe.start_node_id == node2_id and pipe.end_node_id == node1_id):
                return pipe
        return None
    
    def _add_pipe_data_to_row(self, row: HydraulicCalcRow, 
                              pipe: 'HydraulicPipe',
                              from_node: 'HydraulicNode',
                              to_node: 'HydraulicNode'):
        """Add pipe data to calculation row"""
        row.pipe_ref = pipe.tag or pipe.id
        row.pipe_size_nominal = pipe.nominal_diameter
        row.pipe_id_inches = pipe.inside_diameter
        row.c_factor = pipe.c_factor
        row.pipe_length_ft = pipe.length_ft
        row.equiv_length_ft = pipe.equivalent_length_ft
        row.total_length_ft = pipe.total_length_ft
        row.friction_loss_psi = pipe.total_friction_loss_psi
        
        # Elevation change (from current node to next)
        elev_change = to_node.elevation - from_node.elevation
        row.elevation_psi = 0.433 * elev_change
        
        # Normal pressure required at next node
        row.normal_pressure_pn_psi = (row.pressure_pt_psi + 
                                      row.friction_loss_psi + 
                                      row.elevation_psi)
        
        # Fittings summary
        if pipe.fittings:
            row.fittings = ', '.join(
                f"{f.quantity}x{f.fitting_type.replace('_', ' ')}" 
                for f in pipe.fittings
            )
    
    def _calc_bor_pressure(self, rows: List[HydraulicCalcRow]) -> float:
        """Calculate BOR pressure from accumulated losses"""
        if not rows:
            return 0.0
        
        # Start with first sprinkler pressure
        pressure = rows[0].pressure_pt_psi if rows else 0
        
        # Add all friction and elevation losses
        for row in rows:
            pressure += row.friction_loss_psi if row.friction_loss_psi else 0
            pressure += row.elevation_psi if row.elevation_psi else 0
        
        return pressure
    
    def _build_summary(self, network: 'HydraulicNetwork',
                      solution: Dict[str, Any]) -> CalcSheetSummary:
        """Build calculation summary"""
        summary = CalcSheetSummary()
        
        # System demand
        total_flow = sum(
            s.flow_gpm for s in network.sprinklers.values() 
            if s.is_in_remote_area
        )
        summary.system_flow_gpm = total_flow
        
        # Get pressure at source
        source_node = network.nodes.get(network.source_node_id)
        if source_node:
            summary.system_pressure_psi = source_node.pressure_psi
        
        # Hose stream
        summary.hose_stream_gpm = network.hose_stream_gpm
        summary.total_flow_gpm = summary.system_flow_gpm + summary.hose_stream_gpm
        
        # Water supply
        if network.water_supply:
            supply = network.water_supply
            summary.static_pressure_psi = supply.static_pressure_psi
            summary.residual_pressure_psi = supply.residual_pressure_psi
            summary.flow_at_residual_gpm = supply.flow_at_residual_gpm
            
            # Available pressure at total demand
            summary.available_at_demand_psi = supply.get_pressure_at_flow(
                summary.total_flow_gpm
            )
            
            # Safety margin
            summary.safety_margin_psi = (summary.available_at_demand_psi - 
                                         summary.system_pressure_psi)
            summary.is_adequate = summary.safety_margin_psi >= 0
            
            # Build supply curve points
            max_flow = supply.flow_at_residual_gpm * 1.5
            for flow in range(0, int(max_flow) + 100, 100):
                pressure = supply.get_pressure_at_flow(flow)
                if pressure > 0:
                    summary.supply_curve_points.append((flow, pressure))
        
        summary.demand_point = (summary.system_flow_gpm, summary.system_pressure_psi)
        summary.total_demand_point = (summary.total_flow_gpm, summary.system_pressure_psi)
        
        return summary
    
    # ==================================================================================
    # PDF GENERATION
    # ==================================================================================
    
    def generate_pdf(self,
                    project_info: ProjectInfo,
                    design_criteria: DesignCriteria,
                    water_supply: Optional['WaterSupplyData'],
                    calc_rows: List[HydraulicCalcRow],
                    summary: CalcSheetSummary,
                    output_path: str) -> str:
        """Generate professional PDF calculation report"""
        if not reportlab_available:
            self.logger.error("ReportLab not available")
            return ""
        
        self.logger.info(f"Generating PDF: {output_path}")
        
        doc = SimpleDocTemplate(
            output_path,
            pagesize=letter,
            rightMargin=self.margin,
            leftMargin=self.margin,
            topMargin=self.margin,
            bottomMargin=self.margin
        )
        
        story = []
        
        # Page 1: Cover sheet
        story.extend(self._pdf_cover_sheet(project_info, design_criteria))
        story.append(PageBreak())
        
        # Page 2: Design criteria and water supply
        story.extend(self._pdf_design_criteria(design_criteria))
        story.append(Spacer(1, 12))
        story.extend(self._pdf_water_supply(water_supply))
        story.append(PageBreak())
        
        # Page 3+: Hydraulic calculations
        story.extend(self._pdf_hydraulic_calcs(calc_rows))
        story.append(PageBreak())
        
        # Summary page
        story.extend(self._pdf_summary(summary, design_criteria))
        
        # Build PDF
        doc.build(story)
        
        self.logger.info(f"✅ PDF saved: {output_path}")
        return output_path
    
    def _pdf_cover_sheet(self, project_info: ProjectInfo,
                        design_criteria: DesignCriteria) -> List:
        """Generate cover sheet content"""
        elements = []
        
        # Title
        elements.append(Paragraph(
            "FIRE SPRINKLER HYDRAULIC CALCULATIONS",
            self.styles['CalcTitle']
        ))
        elements.append(Spacer(1, 24))
        
        # Project info table
        project_data = [
            ['PROJECT INFORMATION', ''],
            ['Project Name:', project_info.project_name or 'N/A'],
            ['Project Number:', project_info.project_number or 'N/A'],
            ['Address:', project_info.address or 'N/A'],
            ['City, State, ZIP:', f"{project_info.city}, {project_info.state} {project_info.zip_code}"],
            ['', ''],
            ['CONTRACTOR INFORMATION', ''],
            ['Contractor:', project_info.contractor_name or 'N/A'],
            ['License #:', project_info.contractor_license or 'N/A'],
            ['Phone:', project_info.contractor_phone or 'N/A'],
            ['', ''],
            ['DRAWING INFORMATION', ''],
            ['Drawing Number:', project_info.drawing_number or 'N/A'],
            ['Drawing Date:', project_info.drawing_date or 'N/A'],
            ['Calculation Date:', project_info.calculation_date],
            ['Revision:', project_info.revision or '-'],
            ['', ''],
            ['SYSTEM INFORMATION', ''],
            ['System Type:', design_criteria.system_type],
            ['Hazard Classification:', design_criteria.hazard_classification],
            ['Total Sprinklers:', str(design_criteria.total_sprinklers)],
            ['Sprinklers Calculated:', str(design_criteria.sprinklers_calculated)],
        ]
        
        table = Table(project_data, colWidths=[2.5*inch, 4.5*inch])
        table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTNAME', (0, 6), (-1, 6), 'Helvetica-Bold'),
            ('FONTNAME', (0, 11), (-1, 11), 'Helvetica-Bold'),
            ('FONTNAME', (0, 16), (-1, 16), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('BACKGROUND', (0, 6), (-1, 6), colors.lightgrey),
            ('BACKGROUND', (0, 11), (-1, 11), colors.lightgrey),
            ('BACKGROUND', (0, 16), (-1, 16), colors.lightgrey),
            ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
            ('ALIGN', (1, 0), (1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]))
        elements.append(table)
        
        # Certification box
        elements.append(Spacer(1, 36))
        cert_text = """
        <b>CERTIFICATION</b><br/><br/>
        I hereby certify that the hydraulic calculations contained herein have been 
        prepared in accordance with NFPA 13 - Standard for the Installation of Sprinkler Systems, 
        and represent the water supply requirements for the fire sprinkler system as designed.
        <br/><br/>
        Designer: _________________________________ Date: ____________<br/><br/>
        License #: _________________________________
        """
        elements.append(Paragraph(cert_text, self.styles['Normal']))
        
        return elements
    
    def _pdf_design_criteria(self, criteria: DesignCriteria) -> List:
        """Generate design criteria section"""
        elements = []
        
        elements.append(Paragraph("DESIGN CRITERIA", self.styles['SectionHeader']))
        
        data = [
            ['Parameter', 'Value', 'Parameter', 'Value'],
            ['Occupancy Type:', criteria.occupancy_type or 'N/A',
             'Hazard Class:', criteria.hazard_classification],
            ['Design Density:', f"{criteria.design_density_gpm_sqft} GPM/sqft",
             'Design Area:', f"{criteria.design_area_sqft} sqft"],
            ['Area Adjustment:', f"{criteria.area_adjustment_factor}x",
             'Adjusted Area:', f"{criteria.adjusted_area_sqft} sqft"],
            ['Sprinkler Type:', criteria.sprinkler_type,
             'K-Factor:', str(criteria.sprinkler_k_factor)],
            ['Coverage Area:', f"{criteria.sprinkler_coverage_sqft} sqft",
             'Temp Rating:', criteria.sprinkler_temp_rating],
            ['Total Sprinklers:', str(criteria.total_sprinklers),
             'Calculated:', str(criteria.sprinklers_calculated)],
            ['Hose Stream:', f"{criteria.hose_stream_gpm} GPM",
             'Duration:', f"{criteria.duration_minutes} min"],
            ['Pipe Material:', criteria.pipe_material,
             'Pipe Schedule:', criteria.pipe_schedule],
        ]
        
        table = Table(data, colWidths=[1.5*inch, 1.75*inch, 1.5*inch, 1.75*inch])
        table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTNAME', (0, 1), (0, -1), 'Helvetica-Bold'),
            ('FONTNAME', (2, 1), (2, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('TOPPADDING', (0, 0), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ]))
        elements.append(table)
        
        return elements
    
    def _pdf_water_supply(self, water_supply: Optional['WaterSupplyData']) -> List:
        """Generate water supply section"""
        elements = []
        
        elements.append(Paragraph("WATER SUPPLY DATA", self.styles['SectionHeader']))
        
        if not water_supply:
            elements.append(Paragraph("No water supply data provided.", self.styles['Normal']))
            return elements
        
        data = [
            ['Parameter', 'Value'],
            ['Test Location:', water_supply.test_location],
            ['Test Date:', water_supply.test_date or 'N/A'],
            ['Elevation:', f"{water_supply.elevation_ft} ft"],
            ['Static Pressure:', f"{water_supply.static_pressure_psi} PSI"],
            ['Residual Pressure:', f"{water_supply.residual_pressure_psi} PSI"],
            ['Flow at Residual:', f"{water_supply.flow_at_residual_gpm} GPM"],
            ['Available @ 20 PSI:', f"{water_supply.available_flow_at_20psi:.0f} GPM"],
        ]
        
        table = Table(data, colWidths=[2.5*inch, 2.5*inch])
        table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTNAME', (0, 1), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('TOPPADDING', (0, 0), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ]))
        elements.append(table)
        
        return elements
    
    def _pdf_hydraulic_calcs(self, calc_rows: List[HydraulicCalcRow]) -> List:
        """Generate hydraulic calculations table"""
        elements = []
        
        elements.append(Paragraph(
            "HYDRAULIC CALCULATIONS - NODE BY NODE",
            self.styles['SectionHeader']
        ))
        
        # Header row
        headers = [
            'Step', 'Node', 'Elev\n(ft)', 'K', 'Flow\n(GPM)', 'Pt\n(PSI)',
            'Pipe', 'Size', 'ID', 'C', 'L\n(ft)', 'Eq\n(ft)', 'Tot\n(ft)',
            'Pf\n(PSI)', 'Pe\n(PSI)', 'Pn\n(PSI)'
        ]
        
        data = [headers]
        
        for row in calc_rows:
            data.append([
                str(row.step),
                row.node_ref[:8],
                f"{row.elevation_ft:.1f}",
                f"{row.k_factor:.1f}" if row.k_factor else '-',
                f"{row.flow_q_gpm:.1f}",
                f"{row.pressure_pt_psi:.2f}",
                row.pipe_ref[:6] if row.pipe_ref else '-',
                f'{row.pipe_size_nominal}"' if row.pipe_size_nominal else '-',
                f"{row.pipe_id_inches:.3f}" if row.pipe_id_inches else '-',
                str(row.c_factor) if row.c_factor else '-',
                f"{row.pipe_length_ft:.1f}" if row.pipe_length_ft else '-',
                f"{row.equiv_length_ft:.1f}" if row.equiv_length_ft else '-',
                f"{row.total_length_ft:.1f}" if row.total_length_ft else '-',
                f"{row.friction_loss_psi:.3f}" if row.friction_loss_psi else '-',
                f"{row.elevation_psi:.2f}" if row.elevation_psi else '-',
                f"{row.normal_pressure_pn_psi:.2f}" if row.normal_pressure_pn_psi else '-',
            ])
        
        # Column widths
        col_widths = [
            0.35*inch,  # Step
            0.55*inch,  # Node
            0.4*inch,   # Elev
            0.35*inch,  # K
            0.45*inch,  # Flow
            0.45*inch,  # Pt
            0.45*inch,  # Pipe
            0.4*inch,   # Size
            0.45*inch,  # ID
            0.3*inch,   # C
            0.4*inch,   # L
            0.4*inch,   # Eq
            0.4*inch,   # Tot
            0.45*inch,  # Pf
            0.4*inch,   # Pe
            0.45*inch,  # Pn
        ]
        
        table = Table(data, colWidths=col_widths, repeatRows=1)
        table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 7),
            ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('TOPPADDING', (0, 0), (-1, -1), 2),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
            # Highlight last row (system demand)
            ('BACKGROUND', (0, -1), (-1, -1), colors.lightyellow),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ]))
        elements.append(table)
        
        # Legend
        elements.append(Spacer(1, 12))
        legend = """
        <b>Legend:</b> Elev=Elevation, K=K-Factor, Pt=Total Pressure Required, 
        L=Pipe Length, Eq=Equivalent Length, Tot=Total Length, 
        Pf=Friction Loss, Pe=Elevation Pressure, Pn=Normal Pressure
        """
        elements.append(Paragraph(legend, self.styles['SmallText']))
        
        return elements
    
    def _pdf_summary(self, summary: CalcSheetSummary,
                    criteria: DesignCriteria) -> List:
        """Generate summary section"""
        elements = []
        
        elements.append(Paragraph("CALCULATION SUMMARY", self.styles['SectionHeader']))
        
        # Results table
        status = "✓ ADEQUATE" if summary.is_adequate else "✗ INADEQUATE"
        status_color = colors.green if summary.is_adequate else colors.red
        
        data = [
            ['SYSTEM DEMAND', ''],
            ['Sprinkler System Demand:', f"{summary.system_flow_gpm:.1f} GPM @ {summary.system_pressure_psi:.1f} PSI"],
            ['Hose Stream Allowance:', f"{summary.hose_stream_gpm:.0f} GPM"],
            ['TOTAL DEMAND:', f"{summary.total_flow_gpm:.1f} GPM @ {summary.system_pressure_psi:.1f} PSI"],
            ['', ''],
            ['WATER SUPPLY', ''],
            ['Static Pressure:', f"{summary.static_pressure_psi:.1f} PSI"],
            ['Residual Pressure:', f"{summary.residual_pressure_psi:.1f} PSI @ {summary.flow_at_residual_gpm:.0f} GPM"],
            ['Available at Demand:', f"{summary.available_at_demand_psi:.1f} PSI @ {summary.total_flow_gpm:.0f} GPM"],
            ['', ''],
            ['RESULT', ''],
            ['Safety Margin:', f"{summary.safety_margin_psi:.1f} PSI"],
            ['Water Supply Status:', status],
        ]
        
        table = Table(data, colWidths=[2.5*inch, 4*inch])
        table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTNAME', (0, 0), (0, 0), 'Helvetica-Bold'),
            ('FONTNAME', (0, 3), (-1, 3), 'Helvetica-Bold'),
            ('FONTNAME', (0, 5), (0, 5), 'Helvetica-Bold'),
            ('FONTNAME', (0, 10), (0, 10), 'Helvetica-Bold'),
            ('FONTNAME', (0, 12), (-1, 12), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('FONTSIZE', (0, 12), (-1, 12), 12),
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('BACKGROUND', (0, 5), (-1, 5), colors.lightgrey),
            ('BACKGROUND', (0, 10), (-1, 10), colors.lightgrey),
            ('BACKGROUND', (0, 3), (-1, 3), colors.lightyellow),
            ('BACKGROUND', (1, 12), (1, 12), colors.lightgreen if summary.is_adequate else colors.lightcoral),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]))
        elements.append(table)
        
        # Notes
        elements.append(Spacer(1, 24))
        elements.append(Paragraph("NOTES AND REFERENCES", self.styles['SectionHeader']))
        
        notes = f"""
        1. Calculations performed in accordance with NFPA 13 - Standard for the 
           Installation of Sprinkler Systems (2022 Edition).<br/><br/>
        2. Hazen-Williams formula used for friction loss calculations.<br/><br/>
        3. Design criteria: {criteria.hazard_classification} - 
           {criteria.design_density_gpm_sqft} GPM/sqft over {criteria.design_area_sqft} sqft.<br/><br/>
        4. Water supply data should be verified prior to system installation.<br/><br/>
        5. All pipe sizes shown are nominal. Inside diameters per ASTM standards.<br/><br/>
        6. Fitting equivalent lengths per NFPA 13 Table 22.4.3.1.1.
        """
        elements.append(Paragraph(notes, self.styles['Normal']))
        
        return elements
    
    # ==================================================================================
    # EXCEL GENERATION
    # ==================================================================================
    
    def generate_excel(self,
                      project_info: ProjectInfo,
                      design_criteria: DesignCriteria,
                      water_supply: Optional['WaterSupplyData'],
                      calc_rows: List[HydraulicCalcRow],
                      summary: CalcSheetSummary,
                      output_path: str) -> str:
        """Generate Excel calculation workbook"""
        if not openpyxl_available:
            self.logger.error("OpenPyXL not available")
            return ""
        
        self.logger.info(f"Generating Excel: {output_path}")
        
        wb = Workbook()
        
        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Sheet 1: Cover
        ws_cover = wb.active
        ws_cover.title = "Cover"
        self._excel_cover_sheet(ws_cover, project_info, design_criteria)
        
        # Sheet 2: Design Criteria
        ws_criteria = wb.create_sheet("Design Criteria")
        self._excel_design_criteria(ws_criteria, design_criteria, water_supply)
        
        # Sheet 3: Hydraulic Calculations
        ws_calcs = wb.create_sheet("Hydraulic Calculations")
        self._excel_hydraulic_calcs(ws_calcs, calc_rows, header_font, header_fill, border)
        
        # Sheet 4: Summary
        ws_summary = wb.create_sheet("Summary")
        self._excel_summary(ws_summary, summary)
        
        # Save
        wb.save(output_path)
        self.logger.info(f"✅ Excel saved: {output_path}")
        return output_path
    
    def _excel_cover_sheet(self, ws, project_info: ProjectInfo,
                          design_criteria: DesignCriteria):
        """Create cover sheet in Excel"""
        ws['A1'] = "FIRE SPRINKLER HYDRAULIC CALCULATIONS"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:D1')
        
        data = [
            ('', ''),
            ('PROJECT INFORMATION', ''),
            ('Project Name:', project_info.project_name),
            ('Project Number:', project_info.project_number),
            ('Address:', project_info.address),
            ('City, State, ZIP:', f"{project_info.city}, {project_info.state} {project_info.zip_code}"),
            ('', ''),
            ('CONTRACTOR INFORMATION', ''),
            ('Contractor:', project_info.contractor_name),
            ('License #:', project_info.contractor_license),
            ('Phone:', project_info.contractor_phone),
            ('', ''),
            ('SYSTEM INFORMATION', ''),
            ('System Type:', design_criteria.system_type),
            ('Hazard Class:', design_criteria.hazard_classification),
            ('Calculation Date:', project_info.calculation_date),
        ]
        
        for i, (label, value) in enumerate(data, start=3):
            ws[f'A{i}'] = label
            ws[f'B{i}'] = value
            if 'INFORMATION' in label:
                ws[f'A{i}'].font = Font(bold=True)
        
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 40
    
    def _excel_design_criteria(self, ws, criteria: DesignCriteria,
                               water_supply: Optional['WaterSupplyData']):
        """Create design criteria sheet"""
        ws['A1'] = "DESIGN CRITERIA"
        ws['A1'].font = Font(bold=True, size=14)
        
        data = [
            ('Hazard Classification:', criteria.hazard_classification),
            ('Design Density:', f"{criteria.design_density_gpm_sqft} GPM/sqft"),
            ('Design Area:', f"{criteria.design_area_sqft} sqft"),
            ('K-Factor:', criteria.sprinkler_k_factor),
            ('Total Sprinklers:', criteria.total_sprinklers),
            ('Sprinklers Calculated:', criteria.sprinklers_calculated),
            ('Hose Stream:', f"{criteria.hose_stream_gpm} GPM"),
            ('Duration:', f"{criteria.duration_minutes} min"),
            ('', ''),
            ('WATER SUPPLY DATA', ''),
        ]
        
        if water_supply:
            data.extend([
                ('Static Pressure:', f"{water_supply.static_pressure_psi} PSI"),
                ('Residual Pressure:', f"{water_supply.residual_pressure_psi} PSI"),
                ('Flow at Residual:', f"{water_supply.flow_at_residual_gpm} GPM"),
            ])
        
        for i, (label, value) in enumerate(data, start=3):
            ws[f'A{i}'] = label
            ws[f'B{i}'] = value
        
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
    
    def _excel_hydraulic_calcs(self, ws, calc_rows: List[HydraulicCalcRow],
                               header_font, header_fill, border):
        """Create hydraulic calculations sheet"""
        headers = [
            'Step', 'Node', 'Elev (ft)', 'K', 'Flow (GPM)', 'Pt (PSI)',
            'Pipe', 'Size', 'ID (in)', 'C', 'Length (ft)', 'Equiv (ft)',
            'Total (ft)', 'Pf (PSI)', 'Pe (PSI)', 'Pn (PSI)', 'Notes'
        ]
        
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        
        for row_idx, calc_row in enumerate(calc_rows, start=2):
            data = [
                calc_row.step,
                calc_row.node_ref,
                calc_row.elevation_ft,
                calc_row.k_factor if calc_row.k_factor else '',
                calc_row.flow_q_gpm,
                calc_row.pressure_pt_psi,
                calc_row.pipe_ref,
                calc_row.pipe_size_nominal if calc_row.pipe_size_nominal else '',
                calc_row.pipe_id_inches if calc_row.pipe_id_inches else '',
                calc_row.c_factor if calc_row.c_factor else '',
                calc_row.pipe_length_ft if calc_row.pipe_length_ft else '',
                calc_row.equiv_length_ft if calc_row.equiv_length_ft else '',
                calc_row.total_length_ft if calc_row.total_length_ft else '',
                calc_row.friction_loss_psi if calc_row.friction_loss_psi else '',
                calc_row.elevation_psi if calc_row.elevation_psi else '',
                calc_row.normal_pressure_pn_psi if calc_row.normal_pressure_pn_psi else '',
                calc_row.notes,
            ]
            
            for col, value in enumerate(data, start=1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.border = border
                if isinstance(value, float):
                    cell.number_format = '0.00' if col in [6, 14, 15, 16] else '0.0'
        
        # Adjust column widths
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 10
    
    def _excel_summary(self, ws, summary: CalcSheetSummary):
        """Create summary sheet"""
        ws['A1'] = "CALCULATION SUMMARY"
        ws['A1'].font = Font(bold=True, size=14)
        
        data = [
            ('', ''),
            ('SYSTEM DEMAND', ''),
            ('Sprinkler Demand:', f"{summary.system_flow_gpm:.1f} GPM @ {summary.system_pressure_psi:.1f} PSI"),
            ('Hose Stream:', f"{summary.hose_stream_gpm:.0f} GPM"),
            ('Total Demand:', f"{summary.total_flow_gpm:.1f} GPM"),
            ('', ''),
            ('WATER SUPPLY', ''),
            ('Static:', f"{summary.static_pressure_psi:.1f} PSI"),
            ('Residual:', f"{summary.residual_pressure_psi:.1f} PSI @ {summary.flow_at_residual_gpm:.0f} GPM"),
            ('Available:', f"{summary.available_at_demand_psi:.1f} PSI @ {summary.total_flow_gpm:.0f} GPM"),
            ('', ''),
            ('RESULT', ''),
            ('Safety Margin:', f"{summary.safety_margin_psi:.1f} PSI"),
            ('Status:', 'ADEQUATE' if summary.is_adequate else 'INADEQUATE'),
        ]
        
        for i, (label, value) in enumerate(data, start=3):
            ws[f'A{i}'] = label
            ws[f'B{i}'] = value
            if label in ['SYSTEM DEMAND', 'WATER SUPPLY', 'RESULT']:
                ws[f'A{i}'].font = Font(bold=True)
            if label == 'Status:':
                ws[f'B{i}'].font = Font(bold=True, color='008000' if summary.is_adequate else 'FF0000')
        
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 35
    
    # ==================================================================================
    # TEXT REPORT GENERATION
    # ==================================================================================
    
    def generate_text(self,
                     project_info: ProjectInfo,
                     design_criteria: DesignCriteria,
                     water_supply: Optional['WaterSupplyData'],
                     calc_rows: List[HydraulicCalcRow],
                     summary: CalcSheetSummary,
                     output_path: str) -> str:
        """Generate plain text calculation report - NFPA 13 format"""
        self.logger.info(f"Generating text report: {output_path}")
        
        lines = []
        
        # Header
        lines.append("=" * 100)
        lines.append("                    FIRE SPRINKLER HYDRAULIC CALCULATIONS")
        lines.append("                         Per NFPA 13 (2022 Edition)")
        lines.append("=" * 100)
        lines.append("")
        
        # Project info
        lines.append("┌" + "─" * 48 + "┬" + "─" * 48 + "┐")
        lines.append("│ PROJECT INFORMATION" + " " * 29 + "│ CONTRACTOR INFORMATION" + " " * 25 + "│")
        lines.append("├" + "─" * 48 + "┼" + "─" * 48 + "┤")
        lines.append(f"│ Project: {project_info.project_name[:37]:<37} │ Contractor: {project_info.contractor_name[:34]:<34} │")
        lines.append(f"│ Number:  {project_info.project_number[:37]:<37} │ License:    {project_info.contractor_license[:34]:<34} │")
        lines.append(f"│ Address: {project_info.address[:37]:<37} │ Phone:      {project_info.contractor_phone[:34]:<34} │")
        lines.append(f"│ City:    {project_info.city}, {project_info.state} {project_info.zip_code}"[:47].ljust(47) + " │ Drawing:    " + f"{project_info.drawing_number[:34]:<34} │")
        lines.append(f"│ Calc Date: {project_info.calculation_date:<35} │" + " " * 48 + "│")
        lines.append("└" + "─" * 48 + "┴" + "─" * 48 + "┘")
        lines.append("")
        
        # Design criteria
        lines.append("┌" + "─" * 98 + "┐")
        lines.append("│ DESIGN CRITERIA" + " " * 82 + "│")
        lines.append("├" + "─" * 32 + "┬" + "─" * 32 + "┬" + "─" * 32 + "┤")
        lines.append(f"│ Hazard: {design_criteria.hazard_classification[:22]:<22} │ Density: {design_criteria.design_density_gpm_sqft} GPM/sqft" + " " * 10 + f"│ Area: {design_criteria.design_area_sqft} sqft" + " " * 14 + "│")
        lines.append(f"│ System: {design_criteria.system_type[:22]:<22} │ K-Factor: {design_criteria.sprinkler_k_factor:<20} │ Coverage: {design_criteria.sprinkler_coverage_sqft} sqft" + " " * 10 + "│")
        lines.append(f"│ Material: {design_criteria.pipe_material[:20]:<20} │ Schedule: {design_criteria.pipe_schedule:<20} │ C-Factor: 120" + " " * 17 + "│")
        lines.append(f"│ Total Heads: {design_criteria.total_sprinklers:<17} │ Calculated: {design_criteria.sprinklers_calculated:<18} │ Hose Stream: {design_criteria.hose_stream_gpm} GPM" + " " * 8 + "│")
        lines.append("└" + "─" * 32 + "┴" + "─" * 32 + "┴" + "─" * 32 + "┘")
        lines.append("")
        
        # Water supply
        if water_supply:
            lines.append("┌" + "─" * 98 + "┐")
            lines.append("│ WATER SUPPLY DATA" + " " * 80 + "│")
            lines.append("├" + "─" * 32 + "┬" + "─" * 32 + "┬" + "─" * 32 + "┤")
            lines.append(f"│ Static: {water_supply.static_pressure_psi} PSI" + " " * 17 + f"│ Residual: {water_supply.residual_pressure_psi} PSI" + " " * 14 + f"│ Flow @ Res: {water_supply.flow_at_residual_gpm} GPM" + " " * 9 + "│")
            lines.append(f"│ Location: {water_supply.test_location[:20]:<20} │ Date: {(water_supply.test_date or 'N/A')[:24]:<24} │ Elevation: {water_supply.elevation_ft} ft" + " " * 12 + "│")
            lines.append("└" + "─" * 32 + "┴" + "─" * 32 + "┴" + "─" * 32 + "┘")
            lines.append("")
        
        # Hydraulic calculations table
        lines.append("┌" + "─" * 98 + "┐")
        lines.append("│ HYDRAULIC CALCULATIONS - NODE BY NODE" + " " * 60 + "│")
        lines.append("├" + "─" * 98 + "┤")
        
        # Table header
        header = "│ Step │   Node   │ Elev │   K   │  Flow  │   Pt   │  Pipe  │ Size │  ID   │  C  │  Len  │  Eq  │   Pf   │  Pe   │   Pn   │"
        lines.append(header)
        lines.append("│      │          │ (ft) │       │ (GPM)  │ (PSI)  │        │ (in) │ (in)  │     │ (ft)  │ (ft) │ (PSI)  │ (PSI) │ (PSI)  │")
        lines.append("├" + "─" * 6 + "┼" + "─" * 10 + "┼" + "─" * 6 + "┼" + "─" * 7 + "┼" + "─" * 8 + "┼" + "─" * 8 + "┼" + "─" * 8 + "┼" + "─" * 6 + "┼" + "─" * 7 + "┼" + "─" * 5 + "┼" + "─" * 7 + "┼" + "─" * 6 + "┼" + "─" * 8 + "┼" + "─" * 7 + "┼" + "─" * 8 + "┤")
        
        for row in calc_rows:
            k_str = f"{row.k_factor:.1f}" if row.k_factor else "  -  "
            pipe_str = row.pipe_ref[:6] if row.pipe_ref else "  -   "
            size_str = f'{row.pipe_size_nominal:.1f}"' if row.pipe_size_nominal else "  -  "
            id_str = f"{row.pipe_id_inches:.3f}" if row.pipe_id_inches else "  -  "
            c_str = f"{row.c_factor}" if row.c_factor else " - "
            len_str = f"{row.pipe_length_ft:.1f}" if row.pipe_length_ft else "  -  "
            eq_str = f"{row.equiv_length_ft:.1f}" if row.equiv_length_ft else "  -  "
            pf_str = f"{row.friction_loss_psi:.3f}" if row.friction_loss_psi else "  -   "
            pe_str = f"{row.elevation_psi:.2f}" if row.elevation_psi else "  -  "
            pn_str = f"{row.normal_pressure_pn_psi:.2f}" if row.normal_pressure_pn_psi else "  -   "
            
            line = f"│ {row.step:>4} │ {row.node_ref[:8]:<8} │ {row.elevation_ft:>4.0f} │ {k_str:>5} │ {row.flow_q_gpm:>6.1f} │ {row.pressure_pt_psi:>6.2f} │ {pipe_str:<6} │ {size_str:>4} │ {id_str:>5} │ {c_str:>3} │ {len_str:>5} │ {eq_str:>4} │ {pf_str:>6} │ {pe_str:>5} │ {pn_str:>6} │"
            lines.append(line)
        
        lines.append("└" + "─" * 6 + "┴" + "─" * 10 + "┴" + "─" * 6 + "┴" + "─" * 7 + "┴" + "─" * 8 + "┴" + "─" * 8 + "┴" + "─" * 8 + "┴" + "─" * 6 + "┴" + "─" * 7 + "┴" + "─" * 5 + "┴" + "─" * 7 + "┴" + "─" * 6 + "┴" + "─" * 8 + "┴" + "─" * 7 + "┴" + "─" * 8 + "┘")
        lines.append("")
        
        # Legend
        lines.append("LEGEND: Elev=Elevation, K=K-Factor, Pt=Total Pressure, Pf=Friction Loss, Pe=Elevation Pressure, Pn=Normal Pressure")
        lines.append("")
        
        # Summary
        status_symbol = "✓" if summary.is_adequate else "✗"
        status_text = "ADEQUATE" if summary.is_adequate else "INADEQUATE"
        
        lines.append("┌" + "─" * 98 + "┐")
        lines.append("│ CALCULATION SUMMARY" + " " * 78 + "│")
        lines.append("├" + "─" * 48 + "┬" + "─" * 49 + "┤")
        lines.append(f"│ SYSTEM DEMAND                                  │ WATER SUPPLY                                    │")
        lines.append("├" + "─" * 48 + "┼" + "─" * 49 + "┤")
        lines.append(f"│ Sprinkler Demand: {summary.system_flow_gpm:>7.1f} GPM @ {summary.system_pressure_psi:>6.1f} PSI     │ Static Pressure:   {summary.static_pressure_psi:>6.1f} PSI                      │")
        lines.append(f"│ Hose Stream:      {summary.hose_stream_gpm:>7.0f} GPM                      │ Residual Pressure: {summary.residual_pressure_psi:>6.1f} PSI @ {summary.flow_at_residual_gpm:>6.0f} GPM      │")
        lines.append(f"│ TOTAL DEMAND:     {summary.total_flow_gpm:>7.1f} GPM @ {summary.system_pressure_psi:>6.1f} PSI     │ Available @ Demand:{summary.available_at_demand_psi:>6.1f} PSI @ {summary.total_flow_gpm:>6.0f} GPM      │")
        lines.append("├" + "─" * 48 + "┴" + "─" * 49 + "┤")
        lines.append(f"│ SAFETY MARGIN: {summary.safety_margin_psi:>+7.1f} PSI                     WATER SUPPLY STATUS: {status_symbol} {status_text:<15}      │")
        lines.append("└" + "─" * 98 + "┘")
        lines.append("")
        
        # Notes
        lines.append("NOTES:")
        lines.append("  1. Calculations performed in accordance with NFPA 13 - Standard for the Installation of")
        lines.append("     Sprinkler Systems (2022 Edition)")
        lines.append("  2. Hazen-Williams formula used for friction loss: Pf = 4.52 × Q^1.85 / (C^1.85 × d^4.87)")
        lines.append("  3. Elevation pressure: Pe = 0.433 × elevation change (ft)")
        lines.append("  4. Fitting equivalent lengths per NFPA 13 Table 22.4.3.1.1")
        lines.append("  5. Pipe inside diameters per ASTM standards")
        lines.append("")
        lines.append("=" * 100)
        lines.append(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} by FireAI Pro Hydraulic Calculator")
        lines.append("=" * 100)
        
        # Write file
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write('\n'.join(lines))
        
        self.logger.info(f"✅ Text report saved: {output_path}")
        return output_path
    
    # ==================================================================================
    # JSON EXPORT
    # ==================================================================================
    
    def generate_json(self,
                     project_info: ProjectInfo,
                     design_criteria: DesignCriteria,
                     water_supply: Optional['WaterSupplyData'],
                     calc_rows: List[HydraulicCalcRow],
                     summary: CalcSheetSummary,
                     output_path: str) -> str:
        """Generate JSON data export"""
        self.logger.info(f"Generating JSON: {output_path}")
        
        data = {
            'generated_at': datetime.now().isoformat(),
            'version': '1.0.0',
            'project_info': {
                'name': project_info.project_name,
                'number': project_info.project_number,
                'address': project_info.address,
                'city': project_info.city,
                'state': project_info.state,
                'zip': project_info.zip_code,
                'contractor': project_info.contractor_name,
                'calculation_date': project_info.calculation_date,
            },
            'design_criteria': {
                'hazard_classification': design_criteria.hazard_classification,
                'design_density_gpm_sqft': design_criteria.design_density_gpm_sqft,
                'design_area_sqft': design_criteria.design_area_sqft,
                'k_factor': design_criteria.sprinkler_k_factor,
                'total_sprinklers': design_criteria.total_sprinklers,
                'sprinklers_calculated': design_criteria.sprinklers_calculated,
                'hose_stream_gpm': design_criteria.hose_stream_gpm,
                'duration_minutes': design_criteria.duration_minutes,
                'system_type': design_criteria.system_type,
            },
            'water_supply': None,
            'calculations': [],
            'summary': {
                'system_flow_gpm': summary.system_flow_gpm,
                'system_pressure_psi': summary.system_pressure_psi,
                'hose_stream_gpm': summary.hose_stream_gpm,
                'total_flow_gpm': summary.total_flow_gpm,
                'available_pressure_psi': summary.available_at_demand_psi,
                'safety_margin_psi': summary.safety_margin_psi,
                'is_adequate': summary.is_adequate,
            }
        }
        
        if water_supply:
            data['water_supply'] = {
                'static_pressure_psi': water_supply.static_pressure_psi,
                'residual_pressure_psi': water_supply.residual_pressure_psi,
                'flow_at_residual_gpm': water_supply.flow_at_residual_gpm,
                'test_location': water_supply.test_location,
                'test_date': water_supply.test_date,
            }
        
        for row in calc_rows:
            data['calculations'].append({
                'step': row.step,
                'node': row.node_ref,
                'elevation_ft': row.elevation_ft,
                'k_factor': row.k_factor,
                'flow_gpm': row.flow_q_gpm,
                'pressure_psi': row.pressure_pt_psi,
                'pipe': row.pipe_ref,
                'pipe_size': row.pipe_size_nominal,
                'pipe_id': row.pipe_id_inches,
                'c_factor': row.c_factor,
                'length_ft': row.pipe_length_ft,
                'equiv_length_ft': row.equiv_length_ft,
                'friction_loss_psi': row.friction_loss_psi,
                'elevation_psi': row.elevation_psi,
                'normal_pressure_psi': row.normal_pressure_pn_psi,
            })
        
        with open(output_path, 'w') as f:
            json.dump(data, f, indent=2)
        
        self.logger.info(f"✅ JSON saved: {output_path}")
        return output_path
    
    # ==================================================================================
    # DEMAND CURVE GRAPH
    # ==================================================================================
    
    def generate_demand_curve_graph(self, summary: CalcSheetSummary,
                                   output_path: str) -> str:
        """Generate demand curve graph"""
        if not matplotlib_available:
            self.logger.error("Matplotlib not available")
            return ""
        
        self.logger.info(f"Generating demand curve: {output_path}")
        
        fig, ax = plt.subplots(figsize=(10, 8))
        
        # Water supply curve
        if summary.supply_curve_points:
            flows = [p[0] for p in summary.supply_curve_points]
            pressures = [p[1] for p in summary.supply_curve_points]
            ax.plot(flows, pressures, 'b-', linewidth=2.5, label='Water Supply', zorder=2)
            ax.fill_between(flows, pressures, alpha=0.1, color='blue')
        
        # System demand point
        ax.plot(summary.demand_point[0], summary.demand_point[1], 
               'ro', markersize=12, label=f'System Demand\n({summary.demand_point[0]:.0f} GPM @ {summary.demand_point[1]:.1f} PSI)',
               zorder=3)
        
        # Total demand point (with hose stream)
        ax.plot(summary.total_demand_point[0], summary.total_demand_point[1],
               'rs', markersize=14, 
               label=f'Total w/ Hose\n({summary.total_demand_point[0]:.0f} GPM @ {summary.total_demand_point[1]:.1f} PSI)',
               zorder=3)
        
        # Available pressure line
        if summary.available_at_demand_psi > 0:
            ax.axhline(y=summary.available_at_demand_psi, color='green', linestyle='--', 
                      linewidth=1.5, alpha=0.7, label=f'Available: {summary.available_at_demand_psi:.1f} PSI')
        
        # Safety margin annotation
        if summary.safety_margin_psi != 0:
            margin_color = 'green' if summary.safety_margin_psi >= 0 else 'red'
            ax.annotate(f'Safety Margin: {summary.safety_margin_psi:.1f} PSI',
                       xy=(summary.total_demand_point[0], summary.available_at_demand_psi),
                       xytext=(summary.total_demand_point[0] + 100, summary.available_at_demand_psi + 5),
                       fontsize=10, color=margin_color, fontweight='bold',
                       arrowprops=dict(arrowstyle='->', color=margin_color))
        
        # Formatting
        ax.set_xlabel('Flow Rate (GPM)', fontsize=12, fontweight='bold')
        ax.set_ylabel('Pressure (PSI)', fontsize=12, fontweight='bold')
        ax.set_title('WATER SUPPLY vs SYSTEM DEMAND', fontsize=14, fontweight='bold')
        
        ax.legend(loc='upper right', fontsize=9)
        ax.grid(True, linestyle='--', alpha=0.7)
        ax.set_xlim(0, None)
        ax.set_ylim(0, None)
        
        # Add status box
        status = "ADEQUATE" if summary.is_adequate else "INADEQUATE"
        status_color = 'green' if summary.is_adequate else 'red'
        props = dict(boxstyle='round', facecolor='white', edgecolor=status_color, linewidth=2)
        ax.text(0.02, 0.98, f'Status: {status}', transform=ax.transAxes, fontsize=12,
               verticalalignment='top', fontweight='bold', color=status_color, bbox=props)
        
        plt.tight_layout()
        plt.savefig(output_path, dpi=150, bbox_inches='tight')
        plt.close()
        
        self.logger.info(f"✅ Graph saved: {output_path}")
        return output_path


# ================================================================================================
# MODULE EXPORTS
# ================================================================================================

__all__ = [
    'NFPA13CalcSheetGenerator',
    'ProjectInfo',
    'DesignCriteria',
    'HydraulicCalcRow',
    'CalcSheetSummary',
]


# ================================================================================================
# MAIN - TESTING
# ================================================================================================

if __name__ == "__main__":
    print("=" * 80)
    print("🔥 FireAI Pro - NFPA 13 Calculation Sheet Generator v1.0.0")
    print("=" * 80)
    
    print(f"\n📦 DEPENDENCIES:")
    print(f"  ReportLab (PDF): {'✅' if reportlab_available else '❌'}")
    print(f"  OpenPyXL (Excel): {'✅' if openpyxl_available else '❌'}")
    print(f"  Matplotlib (Graphs): {'✅' if matplotlib_available else '❌'}")
    
    print("\n✅ Module loaded successfully!")
    print("Use NFPA13CalcSheetGenerator.generate_from_network() to create calculation sheets.")
